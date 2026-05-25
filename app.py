#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import os
import re
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote

import requests
from dotenv import load_dotenv
from flask import Flask, abort, flash, redirect, render_template, request, url_for

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

load_dotenv()

APP_VERSION = "2026.05.21-r8"
SMARTSHEET_BASE_URL = os.getenv("SMARTSHEET_BASE_URL", "https://api.smartsheet.com/2.0").rstrip("/")
SMARTSHEET_ACCESS_TOKEN = os.getenv("SMARTSHEET_ACCESS_TOKEN", "").strip()
SMARTSHEET_SHEET_ID = os.getenv("SMARTSHEET_SHEET_ID", "").strip()
SMARTSHEET_WEB_URL = os.getenv(
    "SMARTSHEET_WEB_URL",
    "https://app.smartsheet.com/sheets/4mxRwjJvcm57HJ6hwxcgVFhF799qcHM6FxMxw2C1?view=grid&newview=true&filterId=1922378880733060",
).strip()
XLSX_FILE_PATH = os.getenv("XLSX_FILE_PATH", "").strip()
CACHE_TTL_SECONDS = int(os.getenv("CACHE_TTL_SECONDS", "300"))
DEFAULT_RESPONSIBLE = os.getenv("DEFAULT_RESPONSIBLE", "").strip()
ENABLE_SMARTSHEET_WRITE = os.getenv("ENABLE_SMARTSHEET_WRITE", "false").lower() == "true"
PORT = int(os.getenv("PORT", "5000"))

# Column names in the Smartsheet export/API. Keep aliases to tolerate minor rename changes.
COLUMN_ALIASES: Dict[str, List[str]] = {
    "status_checkbox": ["Status"],
    "priority": ["Prioridade"],
    "ticket": ["Ticket"],
    "requester": ["Enviado por", "Solicitante", "Criado Por"],
    "category": ["Categoria"],
    "watchers": ["Observadores"],
    "description": ["Descrição da solicitação ou problema", "Descrição"],
    "description_alt": ["Descrição"],
    "steps": ["Passos de reprodução do problema"],
    "unit": ["Unidade"],
    "cost_center": ["Centro de Custo"],
    "asset_number": ["Número do Patrimônio do Equipamento"],
    "created_at": ["Criado em", "Data real de criação"],
    "created_by": ["Criado Por"],
    "assigned_to": ["Atribuído a"],
    "ticket_status": ["Ticket Status"],
    "solution": ["Solução / Sugestão"],
    "time_spent": ["Tempo Investido"],
    "rating": ["Qual nota você daria para este atendimento?", "Número da nota"],
    "modified_by": ["Modificado por"],
    "modified_at": ["Modificado"],
}

LIST_FIELDS = [
    ("priority", "Prioridade"),
    ("ticket", "Ticket"),
    ("requester", "Enviado por"),
    ("category", "Categoria"),
    ("description", "Descrição da solicitação ou problema"),
    ("steps", "Passos de reprodução do problema"),
    ("unit", "Unidade"),
    ("cost_center", "Centro de Custo"),
    ("asset_number", "Número do Patrimônio do Equipamento"),
    ("created_at", "Criado em"),
]

DETAIL_FIELDS = [
    ("ticket_status", "Ticket Status"),
    ("assigned_to", "Atribuído a"),
    *LIST_FIELDS,
    ("watchers", "Observadores"),
    ("created_by", "Criado por"),
    ("solution", "Solução / Sugestão"),
    ("time_spent", "Tempo Investido"),
    ("rating", "Avaliação"),
    ("modified_by", "Modificado por"),
    ("modified_at", "Modificado"),
]

OPEN_STATUS_WORDS = {
    "em progresso",
    "em andamento",
    "aguardo de retorno",
    "aberto",
    "novo",
    "pendente",
    "a iniciar",
    "aguardando",
}

CLOSED_STATUS_WORDS = {
    "resolvido",
    "concluido",
    "concluído",
    "finalizado",
    "encerrado",
    "cancelado",
    "fechado",
}

UNASSIGNED_WORDS = {
    "",
    "nao atribuido",
    "não atribuído",
    "nao atribuído",
    "não atribuido",
    "unassigned",
    "sem responsavel",
    "sem responsável",
}

app = Flask(__name__)
app.secret_key = os.getenv("APP_SECRET_KEY", "local-dev-secret-change-me")

_sheet_cache: Dict[str, Any] = {"expires_at": 0.0, "tickets": [], "source": "", "error": ""}


# ============================================================
# Helpers
# ============================================================

def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


NON_EMPTY_UNASSIGNED_WORDS = {normalize_text(value) for value in UNASSIGNED_WORDS if normalize_text(value)}

INTERNAL_DEMAND_CATEGORY_PREFIX = normalize_text("TI Demandas Internos do Setor")


def is_internal_demand_category(category: Any) -> bool:
    category_norm = normalize_text(category)
    if not category_norm:
        return False
    return (
        category_norm == INTERNAL_DEMAND_CATEGORY_PREFIX
        or category_norm.startswith(INTERNAL_DEMAND_CATEGORY_PREFIX + " >")
        or category_norm.startswith(INTERNAL_DEMAND_CATEGORY_PREFIX + ">")
    )


def clean_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "Sim" if value else "Não"
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def extract_email(text: str) -> str:
    match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text or "", flags=re.IGNORECASE)
    return match.group(0) if match else ""


def split_contact_text(text: str) -> List[str]:
    text = clean_str(text)
    if not text:
        return []
    parts = re.split(r"\s*(?:,|;|\n|\|)\s*", text)
    return [p.strip() for p in parts if p and p.strip()]


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = clean_str(value)
        if text:
            return text
    return ""


def truncate(text: str, limit: int = 120) -> str:
    text = clean_str(text)
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 1)].rstrip() + "…"


def get_alias_value(cells: Dict[str, Any], key: str) -> str:
    for column_name in COLUMN_ALIASES.get(key, []):
        if column_name in cells:
            value = clean_str(cells.get(column_name))
            if value:
                return value
    return ""


def get_alias_raw(cells: Dict[str, Any], key: str) -> Any:
    for column_name in COLUMN_ALIASES.get(key, []):
        if column_name in cells and cells.get(column_name) is not None:
            return cells.get(column_name)
    return None


def row_has_requester(cells: Dict[str, Any]) -> bool:
    """Só considera ticket real quando o campo Enviado por/Solicitante existe.

    A planilha pode ter linhas vazias, fórmulas ou separadores. Como todo chamado
    válido exige Enviado por, linhas sem esse campo não entram na listagem, filtros
    ou contadores.
    """
    return bool(get_alias_value(cells, "requester"))


def smartsheet_url_for_row(row_id: Optional[str]) -> str:
    # Link de abertura para usuários finais.
    # Mantemos separado do SHEET_ID da API porque nem todos acessam a planilha
    # pelo contexto do detentor/token usado na API.
    if SMARTSHEET_WEB_URL:
        return SMARTSHEET_WEB_URL
    if not SMARTSHEET_SHEET_ID or not row_id:
        return ""
    return f"https://app.smartsheet.com/sheets/{SMARTSHEET_SHEET_ID}?rowId={row_id}"


# ============================================================
# Message generator reused from your current script
# ============================================================

@dataclass
class TicketRow:
    A_ticket: str
    B_col2: str
    C_categoria: str
    D_col4: str
    E_assunto: str
    F_complemento: str
    G_unidade: str
    H_ramal: str
    I_col9: str
    J_data: str
    K_col11: str


def email_only(text: str) -> str:
    email = extract_email(text)
    return email if email else clean_str(text)


def normalized_contacts_for_messages(t: TicketRow) -> Dict[str, str]:
    return {
        "k_raw": clean_str(t.K_col11),
        "k_email": email_only(t.K_col11),
        "d_email": email_only(t.D_col4),
        "b_email": email_only(t.B_col2),
    }


def normalize_url(url: Optional[str]) -> str:
    url = (url or "").strip()
    url = re.sub(r"\s+", "", url)
    return url


def url_for_rocketchat_markdown(url: Optional[str]) -> str:
    url = normalize_url(url)
    if not url:
        return ""
    # Rocket.Chat may interpret ~~ and other markdown characters inside long Smartsheet URLs.
    # Percent-encode only the characters that usually break markdown link parsing.
    replacements = {
        "~": "%7E",
        "(": "%28",
        ")": "%29",
        "[": "%5B",
        "]": "%5D",
    }
    for old_char, new_value in replacements.items():
        url = url.replace(old_char, new_value)
    return url


def markdown_link(label: str, url: Optional[str]) -> str:
    url = url_for_rocketchat_markdown(url)
    if not url:
        return label
    return f"[{label}]({url})"


def survey_block(email: str, survey_link: Optional[str]) -> str:
    email = clean_str(email)
    survey_link = normalize_url(survey_link)

    if survey_link:
        if email:
            email_line = f"O mesmo link também foi enviado para o e-mail {email}, caso prefira acessá-lo por lá."
        else:
            email_line = "O mesmo link também foi enviado para o e-mail informado no chamado, caso prefira acessá-lo por lá."

        link = markdown_link("DESSE LINK", survey_link)
        return (
            f"Se possível, pedimos que avalie este atendimento através do {link}. "
            "A avaliação é rápida e muito importante para que possamos aprimorar continuamente nossos serviços e atendimento.\n\n"
            f"{email_line}\n\n"
            "Agradecemos pelo retorno, confiança e parceria! 😄"
        )

    if email:
        return (
            "Se puder, pedimos também que *avalie este atendimento*: é bem rápido e nos ajuda a melhorar continuamente. "
            f"O link da avaliação foi enviado para o e-mail {email}; por lá você também pode registrar comentários sobre a sua experiência.\n\n"
            "Agradecemos pelo retorno e pela parceria!"
        )

    return (
        "Se puder, pedimos também que *avalie este atendimento*: é bem rápido e nos ajuda a melhorar continuamente.\n\n"
        "Agradecemos pelo retorno e pela parceria!"
    )


def concat_title(t: TicketRow) -> str:
    return f"{t.A_ticket} | {t.C_categoria}-EM ABERTO"


def concat_subject(t: TicketRow) -> str:
    return f"{t.E_assunto} | {t.F_complemento}".strip(" |")


R1_TEXT = "Ola NOME tudo bem? Em que momento posso acessar sua máquina?"
S1_TEXT = "Ola NOME Pode me passar mais detalhes desse chamado, por favor?"

M1_TEXT = (
    "Olá! Espero que esteja tudo bem com você. 😄\n\n"
    "Darei início ao atendimento do chamado:\n"
    "```\n\n"
)
N1_TEXT = (
    "\n``` \n\n"
    "_Importante: O chamado será encerrado automaticamente se não houver resposta em até 3 dias úteis, "
    "após duas tentativas de contato, ou se não forem enviados os recursos ou informações necessárias, "
    "conforme nosso procedimento._"
)


def concat_first_message(t: TicketRow) -> str:
    c = normalized_contacts_for_messages(t)
    return (
        M1_TEXT
        + t.A_ticket
        + " | " + c["k_email"]
        + " | " + t.C_categoria
        + " | " + c["d_email"]
        + " | " + t.E_assunto
        + " | " + t.F_complemento
        + " | " + t.G_unidade
        + " | " + t.H_ramal
        + " | " + t.I_col9
        + " | "
        + N1_TEXT
    )


def concat_conclusion_message(t: TicketRow, survey_link: Optional[str] = None) -> str:
    c = normalized_contacts_for_messages(t)
    return (
        f"Vou encerrar o chamado {t.A_ticket} por aqui. Se precisar de qualquer coisa, é só chamar!\n\n"
        + survey_block(c["k_email"], survey_link)
    )


M3_TEXT = (
    "Olá! Tudo bem? :mascote-mobly:\n\n"
    "Vou iniciar agora o atendimento do seu chamado. A partir deste momento, "
    "*toda a comunicação e atualizações relacionadas a ele devem acontecer exclusivamente na discussão* "
)
N3_TEXT = (
    " , para mantermos o histórico centralizado e facilitar o acompanhamento.\n\n"
    "```"
)
O3_TEXT = (
    "``` \n\n"
    "_Importante: o chamado poderá ser encerrado automaticamente se não houver retorno em até *3 dias úteis*, "
    "após *duas tentativas de contato*, ou caso não sejam enviados os *recursos/informações necessários*, "
    "conforme nosso procedimento._"
)


def concat_opening_info(t: TicketRow, ref: str) -> str:
    c = normalized_contacts_for_messages(t)
    return (
        M3_TEXT
        + ref
        + N3_TEXT
        + " | " + c["k_email"]
        + " | " + t.C_categoria
        + " | " + c["d_email"]
        + " | " + t.E_assunto
        + " | " + t.F_complemento
        + " | " + t.G_unidade
        + " | " + t.H_ramal
        + " | " + t.I_col9
        + " | "
        + O3_TEXT
    )


M4_TEXT = (
    "Olá! :mascote-vita2:\n\n"
    "Estou passando para informar que o chamado "
)
N4_TEXT = (
    " foi *encerrado* e a discussão será *arquivada*, conforme nosso procedimento, pois "
    "*não houve retorno em até 3 dias úteis*, mesmo após *duas tentativas de contato*, e/ou não recebemos as "
    "*informações/recursos necessários* para dar continuidade.\n\n"
    "Se ainda precisar de ajuda, é só nos acionar novamente, estamos à disposição.\n\n"
)


def concat_closure_info(t: TicketRow, ref: str, survey_link: Optional[str] = None) -> str:
    c = normalized_contacts_for_messages(t)
    return M4_TEXT + ref + N4_TEXT + survey_block(c["k_email"], survey_link)


M5_TEXT = (
    "Olá! :mascote-infralouco2:\n\n"
    "O chamado "
)
N5_TEXT = (
    " foi *finalizado* e iremos *arquivar a discussão*. Se surgir qualquer nova dúvida ou necessidade, "
    "fique à vontade para nos acionar novamente, estamos à disposição.\n\n"
)


def concat_conclusion_info(t: TicketRow, ref: str, survey_link: Optional[str] = None) -> str:
    c = normalized_contacts_for_messages(t)
    return M5_TEXT + ref + N5_TEXT + survey_block(c["k_email"], survey_link)


P6_TEXT = (
    "Olá! :mascote-infralouco2:\n\n"
    "Sua solicitação feita pelo chamado "
)
Q6_TEXT = (
    " foi *atendida* e *finalizada*. Se surgir qualquer nova dúvida ou necessidade, "
    "fique à vontade para nos acionar novamente, estamos à disposição.\n\n"
)


def concat_request_conclusion_info(t: TicketRow, ref: str, survey_link: Optional[str] = None) -> str:
    c = normalized_contacts_for_messages(t)
    return P6_TEXT + ref + Q6_TEXT + survey_block(c["k_email"], survey_link)


def build_sections(t: TicketRow, ref_override: Optional[str] = None, survey_link: Optional[str] = None) -> Dict[str, str]:
    ref = clean_str(ref_override) or t.A_ticket
    survey_link = normalize_url(survey_link)
    return {
        "Título": concat_title(t),
        "Assunto": concat_subject(t),
        "Mensagem auxiliar 1": R1_TEXT,
        "Mensagem auxiliar 2": S1_TEXT,
        "Primeira Mensagem": concat_first_message(t),
        "Mensagem de Conclusão": concat_conclusion_message(t, survey_link),
        "Informação de Abertura": concat_opening_info(t, ref),
        "Informação de Encerramento": concat_closure_info(t, ref, survey_link),
        "Informação de Conclusão": concat_conclusion_info(t, ref, survey_link),
        "Informação de Conclusão da Solicitação": concat_request_conclusion_info(t, ref, survey_link),
    }


def build_full_text(sections: Dict[str, str]) -> str:
    def sep(title: str) -> str:
        label = f" {title} "
        width = max(100, len(label) + 10)
        remaining = width - len(label)
        return "_" * (remaining // 2) + label + "_" * (remaining - remaining // 2)

    parts: List[str] = []
    for key, value in sections.items():
        parts.extend([sep(key), value, ""])
    parts.append(sep("FIM"))
    return "\n".join(parts).strip() + "\n"


# ============================================================
# Ticket model and data sources
# ============================================================

@dataclass
class Ticket:
    uid: str
    row_id: Optional[str]
    row_number: Optional[int]
    source: str
    cells: Dict[str, Any]
    assigned_contacts: List[Dict[str, str]] = field(default_factory=list)
    created_index: int = 0

    def v(self, key: str) -> str:
        if key == "description":
            return first_non_empty(get_alias_value(self.cells, "description"), get_alias_value(self.cells, "description_alt"))
        return get_alias_value(self.cells, key)

    @property
    def ticket_number(self) -> str:
        return self.v("ticket") or f"Linha {self.row_number or self.uid}"

    @property
    def requester(self) -> str:
        return self.v("requester")

    @property
    def requester_email(self) -> str:
        return extract_email(self.requester) or self.requester

    @property
    def assigned_to(self) -> str:
        # Exibe exatamente o que veio da coluna quando houver valor.
        # Campo vazio fica vazio para não ser confundido com "Não atribuído".
        assigned = self.v("assigned_to")
        if assigned:
            return assigned
        if self.assigned_contacts:
            values = [c.get("email") or c.get("name") or "" for c in self.assigned_contacts if c]
            values = [value for value in values if value]
            if values:
                return ", ".join(values)
        return ""

    @property
    def is_unassigned(self) -> bool:
        # Regra atual: somente conta como "Não atribuído" quando o campo foi
        # preenchido explicitamente com esse texto ou sinônimo. Campo vazio não conta.
        assigned = normalize_text(self.assigned_to)
        return bool(assigned) and assigned in NON_EMPTY_UNASSIGNED_WORDS

    @property
    def is_attention_unassigned(self) -> bool:
        # Chamado aberto marcado explicitamente como "Não atribuído" deve aparecer no topo
        # para qualquer responsável selecionado.
        return self.is_unassigned and not self.is_closed

    @property
    def status(self) -> str:
        return self.v("ticket_status") or self.v("status_checkbox")

    @property
    def is_closed(self) -> bool:
        status = normalize_text(self.status)
        return any(word in status for word in CLOSED_STATUS_WORDS)

    @property
    def priority(self) -> str:
        return self.v("priority")

    @property
    def is_internal_demand(self) -> bool:
        return is_internal_demand_category(self.v("category"))

    @property
    def demand_type_label(self) -> str:
        return "Interna" if self.is_internal_demand else "Externa"

    def assigned_keys(self) -> List[str]:
        keys: List[str] = []
        for contact in self.assigned_contacts:
            for field_name in ("email", "name"):
                value = normalize_text(contact.get(field_name, ""))
                if value:
                    keys.append(value)
        for part in split_contact_text(self.assigned_to):
            keys.append(normalize_text(part))
            email = extract_email(part)
            if email:
                keys.append(normalize_text(email))
        assigned = normalize_text(self.assigned_to)
        if assigned:
            keys.append(assigned)
        # de-duplicate preserving order
        seen = set()
        result = []
        for key in keys:
            if key and key not in seen:
                result.append(key)
                seen.add(key)
        return result

    def matches_responsible(self, responsible: str) -> bool:
        responsible = responsible or "__all__"
        if responsible == "__all__":
            return True
        if responsible == "__unassigned__":
            return self.is_unassigned
        needle = normalize_text(responsible)
        if not needle:
            return True
        return any(needle == key or needle in key or key in needle for key in self.assigned_keys())

    def matches_any_responsible(self, responsibles: List[str]) -> bool:
        selected = [r for r in responsibles if r]
        if not selected or "__all__" in selected:
            return True
        return any(self.matches_responsible(responsible) for responsible in selected)

    def to_message_row(self) -> TicketRow:
        requester = self.requester_email or self.requester
        return TicketRow(
            A_ticket=self.ticket_number,
            B_col2=requester,
            C_categoria=self.v("category"),
            D_col4=requester,
            E_assunto=self.v("description"),
            F_complemento=self.v("steps"),
            G_unidade=self.v("unit"),
            H_ramal=self.v("asset_number"),
            I_col9=self.v("cost_center"),
            J_data=self.v("created_at"),
            K_col11=requester,
        )

    def list_fields(self) -> List[Tuple[str, str, str, bool]]:
        output = []
        for key, label in LIST_FIELDS:
            value = self.v(key)
            is_long = key in {"description", "steps"} and len(value) > 80
            output.append((key, label, value, is_long))
        return output

    def detail_fields(self) -> List[Tuple[str, str]]:
        output = []
        seen = set()
        for key, label in DETAIL_FIELDS:
            if key in seen:
                continue
            seen.add(key)
            value = self.v(key)
            if value:
                output.append((label, value))
        return output

    def details_json(self) -> str:
        return json.dumps(self.detail_fields(), ensure_ascii=False)


class SmartsheetClient:
    def __init__(self, token: str, sheet_id: str):
        self.token = token
        self.sheet_id = sheet_id
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        })

    def _request(self, method: str, path: str, **kwargs: Any) -> Any:
        url = f"{SMARTSHEET_BASE_URL}{path}"
        response = self.session.request(method, url, timeout=30, **kwargs)
        if not response.ok:
            try:
                detail = response.json()
            except Exception:
                detail = response.text
            raise RuntimeError(f"Erro Smartsheet {response.status_code}: {detail}")
        return response.json()

    def get_sheet(self) -> Dict[str, Any]:
        # include=objectValue is important for contact/multi-contact columns.
        return self._request("GET", f"/sheets/{self.sheet_id}", params={"include": "objectValue"})

    def update_row_cells(self, row_id: str, column_updates: Dict[str, Any], column_map: Dict[str, int]) -> Any:
        if not ENABLE_SMARTSHEET_WRITE:
            raise RuntimeError("Escrita no Smartsheet está desativada. Configure ENABLE_SMARTSHEET_WRITE=true para habilitar.")
        cells = []
        for column_name, value in column_updates.items():
            column_id = column_map.get(column_name)
            if column_id:
                cells.append({"columnId": column_id, "value": value})
        if not cells:
            raise RuntimeError("Nenhuma coluna válida para atualizar.")
        payload = [{"id": int(row_id), "cells": cells}]
        return self._request("PUT", f"/sheets/{self.sheet_id}/rows", json=payload)


def cell_display_from_api(cell: Dict[str, Any]) -> str:
    if cell.get("displayValue") not in (None, ""):
        return clean_str(cell.get("displayValue"))

    object_value = cell.get("objectValue")
    if isinstance(object_value, dict):
        values = object_value.get("values") or object_value.get("contacts") or []
        if isinstance(values, list) and values:
            names = []
            for contact in values:
                if not isinstance(contact, dict):
                    continue
                name = contact.get("name") or contact.get("displayName") or ""
                email = contact.get("email") or ""
                names.append(email or name)
            if names:
                return ", ".join(names)
        email = object_value.get("email")
        name = object_value.get("name") or object_value.get("displayName")
        if email or name:
            return email or name

    if cell.get("value") not in (None, ""):
        return clean_str(cell.get("value"))
    return ""


def contacts_from_api_cell(cell: Dict[str, Any]) -> List[Dict[str, str]]:
    contacts: List[Dict[str, str]] = []
    object_value = cell.get("objectValue")
    if isinstance(object_value, dict):
        values = object_value.get("values") or object_value.get("contacts") or []
        if isinstance(values, list):
            for contact in values:
                if isinstance(contact, dict):
                    email = clean_str(contact.get("email"))
                    name = clean_str(contact.get("name") or contact.get("displayName"))
                    if email or name:
                        contacts.append({"email": email, "name": name})
        else:
            email = clean_str(object_value.get("email"))
            name = clean_str(object_value.get("name") or object_value.get("displayName"))
            if email or name:
                contacts.append({"email": email, "name": name})

    if not contacts:
        text = cell_display_from_api(cell)
        for part in split_contact_text(text):
            email = extract_email(part)
            contacts.append({"email": email or part, "name": part if not email else ""})
    return contacts


def load_from_smartsheet() -> Tuple[List[Ticket], str]:
    client = SmartsheetClient(SMARTSHEET_ACCESS_TOKEN, SMARTSHEET_SHEET_ID)
    sheet = client.get_sheet()
    columns = {str(col.get("id")): col.get("title") for col in sheet.get("columns", [])}
    assigned_column_ids = {
        str(col.get("id")) for col in sheet.get("columns", []) if col.get("title") in COLUMN_ALIASES["assigned_to"]
    }

    tickets: List[Ticket] = []
    for index, row in enumerate(sheet.get("rows", []), start=1):
        cells: Dict[str, Any] = {}
        assigned_contacts: List[Dict[str, str]] = []
        for cell in row.get("cells", []):
            column_id = str(cell.get("columnId"))
            column_title = columns.get(column_id)
            if not column_title:
                continue
            cells[column_title] = cell_display_from_api(cell)
            if column_id in assigned_column_ids:
                assigned_contacts = contacts_from_api_cell(cell)

        # Linhas sem Enviado por/Solicitante não são tickets reais.
        if not row_has_requester(cells):
            continue

        uid = f"smartsheet-{row.get('id')}"
        tickets.append(Ticket(
            uid=uid,
            row_id=str(row.get("id")),
            row_number=row.get("rowNumber"),
            source="Smartsheet API",
            cells=cells,
            assigned_contacts=assigned_contacts,
            created_index=index,
        ))
    return tickets, "Smartsheet API"


def load_from_xlsx(path: str) -> Tuple[List[Ticket], str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl não está instalado. Instale as dependências com pip install -r requirements.txt.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [clean_str(value) for value in next(rows_iter)]
    tickets: List[Ticket] = []
    for row_index, row in enumerate(rows_iter, start=2):
        cells = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        # Ignore completely blank rows and non-ticket rows. Every valid ticket
        # must have Enviado por/Solicitante filled.
        if not any(clean_str(value) for value in cells.values()):
            continue
        if not row_has_requester(cells):
            continue
        assigned_text = get_alias_value(cells, "assigned_to")
        contacts = []
        for part in split_contact_text(assigned_text):
            email = extract_email(part)
            contacts.append({"email": email or part, "name": part if not email else ""})
        tickets.append(Ticket(
            uid=f"xlsx-{row_index}",
            row_id=None,
            row_number=row_index,
            source="Arquivo XLSX local",
            cells=cells,
            assigned_contacts=contacts,
            created_index=row_index,
        ))
    return tickets, "Arquivo XLSX local"


def load_tickets(force_refresh: bool = False, allow_stale_cache: bool = False) -> Tuple[List[Ticket], str, str]:
    now = time.time()
    if not force_refresh and _sheet_cache["tickets"] and allow_stale_cache:
        return _sheet_cache["tickets"], _sheet_cache["source"], _sheet_cache["error"]
    if not force_refresh and _sheet_cache["expires_at"] > now:
        return _sheet_cache["tickets"], _sheet_cache["source"], _sheet_cache["error"]

    try:
        if SMARTSHEET_ACCESS_TOKEN and SMARTSHEET_SHEET_ID:
            tickets, source = load_from_smartsheet()
            error = ""
        elif XLSX_FILE_PATH:
            tickets, source = load_from_xlsx(XLSX_FILE_PATH)
            error = ""
        else:
            tickets, source = [], "Sem fonte configurada"
            error = "Configure SMARTSHEET_ACCESS_TOKEN + SMARTSHEET_SHEET_ID ou XLSX_FILE_PATH no arquivo .env / Render."
    except Exception as exc:
        tickets, source, error = [], "Erro", str(exc)

    _sheet_cache.update({
        "tickets": tickets,
        "source": source,
        "error": error,
        "expires_at": now + CACHE_TTL_SECONDS,
    })
    return tickets, source, error


def responsible_options(tickets: Iterable[Ticket]) -> List[Tuple[str, str]]:
    options: Dict[str, str] = {"__all__": "Todos", "__unassigned__": "Não atribuído"}
    for ticket in tickets:
        if ticket.is_unassigned:
            continue
        for contact in ticket.assigned_contacts:
            label = contact.get("email") or contact.get("name") or ""
            if label:
                options.setdefault(label, label)
        if not ticket.assigned_contacts and ticket.assigned_to:
            for part in split_contact_text(ticket.assigned_to):
                if normalize_text(part) not in {normalize_text(x) for x in UNASSIGNED_WORDS}:
                    options.setdefault(part, part)
    ordered = [options["__all__"], options["__unassigned__"]]
    rest = sorted([v for k, v in options.items() if k not in {"__all__", "__unassigned__"}], key=normalize_text)
    return [("__all__", ordered[0]), ("__unassigned__", ordered[1])] + [(value, value) for value in rest]


def filter_tickets(
    tickets: List[Ticket],
    responsibles: List[str],
    status_filter: str,
    demand_type: str,
    search: str,
) -> List[Ticket]:
    search_norm = normalize_text(search)
    result: List[Ticket] = []
    for ticket in tickets:
        if status_filter == "open" and ticket.is_closed:
            continue
        if status_filter == "closed" and not ticket.is_closed:
            continue
        if demand_type == "internal" and not ticket.is_internal_demand:
            continue
        if demand_type == "external" and ticket.is_internal_demand:
            continue
        if search_norm:
            haystack = normalize_text(" ".join([
                ticket.ticket_number,
                ticket.requester,
                ticket.assigned_to,
                ticket.status,
                ticket.v("priority"),
                ticket.v("category"),
                ticket.v("description"),
                ticket.v("steps"),
                ticket.v("unit"),
                ticket.v("cost_center"),
                ticket.v("asset_number"),
            ]))
            if search_norm not in haystack:
                continue

        # Chamados marcados explicitamente como "Não atribuído" entram no topo para
        # qualquer responsável selecionado. Campo vazio não entra nessa regra.
        if ticket.is_attention_unassigned:
            result.append(ticket)
            continue

        if not ticket.matches_any_responsible(responsibles):
            continue
        result.append(ticket)

    # Primeiro os chamados abertos marcados como "Não atribuído", depois a ordem original.
    result.sort(key=lambda item: (0 if item.is_attention_unassigned else 1, item.created_index))
    return result


def get_ticket_or_404(uid: str) -> Ticket:
    # Na tela de mensagens, reaproveita o cache mesmo vencido para evitar
    # reler a planilha a cada clique em Aplicar/copiar/voltar.
    tickets, _source, _error = load_tickets(force_refresh=False, allow_stale_cache=True)
    for ticket in tickets:
        if ticket.uid == uid:
            return ticket
    # Retry once in case the cache expired or the row was not loaded.
    tickets, _source, _error = load_tickets(force_refresh=True)
    for ticket in tickets:
        if ticket.uid == uid:
            return ticket
    abort(404)


def normalize_responsible_filter_value(value: str) -> str:
    """Normaliza valores vindos da URL/.env para as chaves internas do filtro."""
    raw = (value or "").strip()
    norm = normalize_text(raw)
    if norm in {"todos", "todo", "all", "__all__"}:
        return "__all__"
    if norm in {"nao atribuido", "nao atribuidos", "não atribuído", "não atribuídos", "unassigned", "__unassigned__"}:
        return "__unassigned__"
    return raw


def get_selected_responsibles() -> List[str]:
    values = [
        normalize_responsible_filter_value(value)
        for value in request.args.getlist("responsible")
        if value and value.strip()
    ]
    if not values:
        values = [
            normalize_responsible_filter_value(value)
            for value in split_contact_text(DEFAULT_RESPONSIBLE)
            if value and value.strip()
        ] or ["__all__"]
    if "__all__" in values:
        return ["__all__"]
    # De-duplicate preserving order.
    seen = set()
    result: List[str] = []
    for value in values:
        if value and value not in seen:
            result.append(value)
            seen.add(value)
    return result or ["__all__"]


# ============================================================
# Routes
# ============================================================

@app.get("/")
def index():
    force = request.args.get("refresh") == "1"
    tickets, source, error = load_tickets(force_refresh=force)
    if force:
        flash("Dados recarregados.")

    responsibles = get_selected_responsibles()
    status_filter = request.args.get("status", "open")
    demand_type = request.args.get("demand_type", "all")
    if demand_type not in {"all", "internal", "external"}:
        demand_type = "all"
    search = request.args.get("q", "")

    filtered = filter_tickets(tickets, responsibles, status_filter, demand_type, search)

    try:
        page = max(1, int(request.args.get("page", "1")))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get("per_page", "100"))
    except ValueError:
        per_page = 100
    per_page = min(max(per_page, 25), 250)
    total_filtered = len(filtered)
    total_pages = max(1, (total_filtered + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * per_page
    end = start + per_page
    page_tickets = filtered[start:end]

    open_tickets = [ticket for ticket in tickets if not ticket.is_closed]
    options = responsible_options(open_tickets)
    option_keys = [key for key, _label in options]
    for responsible in responsibles:
        if responsible not in option_keys and responsible != "__all__":
            options.append((responsible, responsible))
            option_keys.append(responsible)

    stats = {
        "total": len(tickets),
        "open": len(open_tickets),
        "filtered": len(filtered),
        "unassigned": sum(1 for t in open_tickets if t.is_unassigned),
        "closed": sum(1 for t in tickets if t.is_closed),
        "internal_open": sum(1 for t in open_tickets if t.is_internal_demand),
        "external_open": sum(1 for t in open_tickets if not t.is_internal_demand),
    }
    return_to = request.full_path.rstrip("?")
    return render_template(
        "index.html",
        app_version=APP_VERSION,
        source=source,
        error=error,
        tickets=page_tickets,
        options=options,
        responsibles=responsibles,
        status_filter=status_filter,
        demand_type=demand_type,
        search=search,
        stats=stats,
        pagination={"page": page, "per_page": per_page, "total_pages": total_pages, "total_filtered": total_filtered},
        list_fields=LIST_FIELDS,
        enable_write=ENABLE_SMARTSHEET_WRITE,
        return_to=return_to,
        alert_tickets=[{"uid": t.uid, "ticket": t.ticket_number} for t in filtered if not t.is_closed],
        alert_filter_key="|".join([",".join(responsibles), status_filter, demand_type, normalize_text(search)]),
    )


@app.route("/ticket/<uid>", methods=["GET", "POST"])
def ticket_messages(uid: str):
    ticket = get_ticket_or_404(uid)
    return_to = request.values.get("return_to") or url_for("index")
    # Evita redirecionar para fora do próprio app caso alguém altere a URL manualmente.
    if not return_to.startswith("/"):
        return_to = url_for("index")

    if request.method == "POST":
        reference = request.form.get("reference", "")
        survey_link = request.form.get("survey_link", "")
        return redirect(url_for(
            "ticket_messages",
            uid=uid,
            reference=reference,
            survey_link=survey_link,
            return_to=return_to,
        ))

    reference = request.args.get("reference", "") or ticket.ticket_number
    survey_link = request.args.get("survey_link", "")
    message_row = ticket.to_message_row()
    sections = build_sections(message_row, reference, survey_link)
    full_text = build_full_text(sections)

    return render_template(
        "ticket.html",
        app_version=APP_VERSION,
        ticket=ticket,
        sections=sections,
        full_text=full_text,
        reference=reference,
        survey_link=survey_link,
        return_to=return_to,
        smartsheet_url=smartsheet_url_for_row(ticket.row_id),
    )


@app.get("/healthz")
def healthz():
    tickets, source, error = load_tickets(force_refresh=False)
    status = 200 if not error else 503
    return {"ok": not bool(error), "source": source, "tickets": len(tickets), "error": error}, status


@app.template_filter("short")
def short_filter(text: str, limit: int = 110) -> str:
    return truncate(text, limit)


@app.template_filter("json_details")
def json_details_filter(ticket: Ticket) -> str:
    return ticket.details_json()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT, debug=os.getenv("FLASK_DEBUG", "0") == "1")
